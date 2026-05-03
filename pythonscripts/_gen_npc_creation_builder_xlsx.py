"""
Generate an NPC creation workbook from config/npcdesc.cfg.

Output: npc_creation_builder.xlsx

Workbook tabs:
- Builder: grouped inputs + generated NpcTemplate text block
- FieldDictionary: description of fields and CProps
- SpellLists: spells available to NPCs (from spells.cfg files + npcdesc usage)
- AllVariables: exhaustive variable inventory with observed counts/order
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
NPCDESC_PATH = ROOT / "config" / "npcdesc.cfg"
OUT_PATH = ROOT / "npc_creation_builder.xlsx"


def strip_inline_comment(line: str) -> str:
    # Keep parser simple and consistent with existing repo scripts.
    return line.split("//", 1)[0].rstrip()


def parse_npcdesc(path: Path) -> dict:
    templates: list[dict] = []
    in_template = False
    brace_depth = 0
    current_name = ""
    field_index = 0

    key_counts: Counter[str] = Counter()
    key_variants: defaultdict[str, Counter[str]] = defaultdict(Counter)
    key_pos_sum: defaultdict[str, int] = defaultdict(int)

    cprop_counts: Counter[str] = Counter()
    cprop_variants: defaultdict[str, Counter[str]] = defaultdict(Counter)
    cprop_pos_sum: defaultdict[str, int] = defaultdict(int)

    npc_spells: Counter[str] = Counter()
    spell_variants: defaultdict[str, Counter[str]] = defaultdict(Counter)

    with path.open("r", encoding="utf-8", errors="replace") as f:
        for raw in f:
            line = strip_inline_comment(raw).strip()
            if not line:
                continue
            if line.startswith("#"):
                continue

            m = re.match(r"^NpcTemplate\s+(\S+)", line, flags=re.IGNORECASE)
            if m:
                in_template = True
                brace_depth = 0
                current_name = m.group(1)
                field_index = 0
                templates.append({"name": current_name, "fields": []})
                continue

            if not in_template:
                continue

            if line == "{":
                brace_depth += 1
                continue

            if line == "}":
                brace_depth -= 1
                if brace_depth <= 0:
                    in_template = False
                    current_name = ""
                continue

            parts = line.split()
            if not parts:
                continue

            key = parts[0]
            value = line[len(key) :].strip() if len(line) > len(key) else ""

            templates[-1]["fields"].append((key, value, line))

            k_norm = key.lower()
            key_counts[k_norm] += 1
            key_variants[k_norm][key] += 1
            key_pos_sum[k_norm] += field_index

            if k_norm == "cprop" and len(parts) >= 3:
                cprop = parts[1]
                c_norm = cprop.lower()
                cprop_counts[c_norm] += 1
                cprop_variants[c_norm][cprop] += 1
                cprop_pos_sum[c_norm] += field_index
            elif k_norm == "spell" and len(parts) >= 2:
                sp = parts[1]
                s_norm = sp.lower()
                npc_spells[s_norm] += 1
                spell_variants[s_norm][sp] += 1

            field_index += 1

    return {
        "templates": templates,
        "key_counts": key_counts,
        "key_variants": key_variants,
        "key_pos_sum": key_pos_sum,
        "cprop_counts": cprop_counts,
        "cprop_variants": cprop_variants,
        "cprop_pos_sum": cprop_pos_sum,
        "npc_spells": npc_spells,
        "spell_variants": spell_variants,
    }


def most_common_variant(counter_map: Counter[str]) -> str:
    return max(counter_map.items(), key=lambda kv: (kv[1], -len(kv[0]), kv[0]))[0]


def collect_spell_catalog(root: Path) -> list[dict[str, str]]:
    """
    Parse every spells.cfg and return spell records with numeric id, script, and name.
    """
    records: list[dict[str, str]] = []
    spell_files = sorted(root.rglob("spells.cfg"))
    re_spell = re.compile(r"^\s*Spell\s+(\S+)", flags=re.IGNORECASE)
    re_spell_id = re.compile(r"^\s*SpellId\s+(\S+)", flags=re.IGNORECASE)
    re_name = re.compile(r"^\s*Name\s+(.+)$", flags=re.IGNORECASE)
    re_script = re.compile(r"^\s*Script\s+(\S+)", flags=re.IGNORECASE)

    for path in spell_files:
        current: dict[str, str] | None = None
        try:
            with path.open("r", encoding="utf-8", errors="replace") as f:
                for raw in f:
                    line = strip_inline_comment(raw).strip()
                    if not line:
                        continue

                    m_spell = re_spell.match(line)
                    if m_spell:
                        if current:
                            records.append(current)
                        current = {
                            "header": m_spell.group(1).strip(),
                            "id": "",
                            "script": "",
                            "name": "",
                            "file": str(path.relative_to(root)).replace('\\\\', '/'),
                        }
                        continue

                    if current is None:
                        continue

                    if line == "}":
                        records.append(current)
                        current = None
                        continue

                    m_id = re_spell_id.match(line)
                    if m_id:
                        current["id"] = m_id.group(1).strip()
                        continue

                    m_script = re_script.match(line)
                    if m_script:
                        current["script"] = m_script.group(1).strip()
                        continue

                    m_name = re_name.match(line)
                    if m_name:
                        current["name"] = m_name.group(1).strip()
                        continue
        except OSError:
            continue

    return records


def build_descriptions() -> dict[str, str]:
    return {
        "NpcTemplateID": "Builder-only template id written after NpcTemplate. Use letters/numbers and no spaces.",
        "Name": "Display name for NPC; can use <random> tokens where supported.",
        "script": "AI/behavior script attached to NPC.",
        "objtype": "Body/object type ID (hex or decimal).",
        "Color": "Hue override for body/equipment visuals.",
        "TrueColor": "Permanent hue that should persist through updates.",
        "Gender": "0/1 gender selector used by animation/equipment rules.",
        "Equip": "Equipment template applied on spawn.",
        "STR": "Base strength stat. If HITS is blank and CustomHitsLevel is not set, HP often ends up tracking strength after BaseStrmod is applied.",
        "INT": "Base intelligence stat.",
        "DEX": "Base dexterity stat.",
        "HITS": "Direct hit-point assignment from npcdesc. Use this when you want a fixed HP value instead of strength-derived HP.",
        "MANA": "Base mana pool.",
        "STAM": "Base stamina pool.",
        "alignment": "Behavior allegiance (for example good/evil).",
        "hostile": "If set, NPC auto-hostiles players/targets.",
        "virtue": "Virtue/karma subsystem bucket value used by shard scripts.",
        "Karma": "Karma value used for notoriety/kill rewards.",
        "Fame": "Fame value used for notoriety/kill rewards.",
        "lootgroup": "Loot table group ID.",
        "MagicItemChance": "Chance percent for magic drop rolls.",
        "MagicItemLevel": "Tier for generated magic item quality.",
        "spell": "Spell entry. Repeat this key for multiple spells.",
        "cast_pct": "Caster AI cast probability.",
        "num_casts": "Caster AI cast attempt count.",
        "MoveMode": "Movement flags (L/A/S etc.) for pathing capability.",
        "RunSpeed": "Core run speed tuning field.",
        "dstart": "AI perception/engagement start distance.",
        "saywords": "Enables NPC speech lines during behavior.",
        "speech": "Speech profile/group index.",
        "provoke": "Provocation difficulty/value.",
        "AttackAttribute": "Attack profile label. Kept mainly for staff/Seer info display; shard scripts do not use it to drive combat logic.",
        "AttackSpeed": "Attack speed note field. Kept for template documentation/display; shard scripts do not read this value for combat behavior.",
        "AttackDamage": "Attack damage note field. Kept mainly for staff/Seer info display rather than live combat calculation.",
        "deathsnd": "Death sound ID.",
        "tameskill": "Required animal taming skill.",
        "food": "Food type accepted by tame/feeding logic.",
        "mount": "Mount body/hue or mount setup fields.",
        "Privs": "Privilege flags (often invul for vendors/town NPCs).",
        "Settings": "NPC settings flags (often invul).",
        "guardignore": "If set, guards ignore this NPC.",
        "noloot": "Suppresses corpse loot generation.",
        "title": "Title behavior/style control in name rendering.",
        "dress": "Dress profile for random outfitting.",
        "AR": "Armor rating baseline.",
        "CProp.Type": "Slayer/system classification (e.g., sHuman, sDragonkin).",
        "CProp.BaseStrmod": "Strength modifier applied after base STR. This usually raises derived HP if HP is strength-based.",
        "CProp.BaseIntmod": "Additional intelligence modifier from custom property.",
        "CProp.BaseDexmod": "Additional dexterity modifier from custom property.",
        "CProp.CustomHitsLevel": "Custom HP override used by shard scripts. Common shard convention stores HP in hundredths, so i30000 becomes 300 HP.",
        "CProp.PermPoisonImmunity": "Permanent poison immunity tier.",
        "CProp.PermMagicImmunity": "Permanent magic immunity tier.",
        "CProp.Permmr": "Permanent magic resistance tier/value.",
        "CProp.FireProtection": "Fire resistance/protection value.",
        "CProp.WaterProtection": "Water resistance/protection value.",
        "CProp.AirProtection": "Air resistance/protection value.",
        "CProp.EarthProtection": "Earth resistance/protection value.",
        "CProp.NecroProtection": "Necro resistance/protection value.",
        "CProp.HolyProtection": "Holy resistance/protection value.",
        "CProp.PhysicalProtection": "Physical resistance/protection value.",
        "CProp.BaseHpRegen": "Passive HP regeneration modifier.",
        "CProp.BaseManaRegen": "Passive mana regeneration modifier.",
        "skill_name": "Skill key name. Pick from dropdown (e.g. Magery, Tactics).",
        "skill_value": "Skill numeric value for the selected skill key.",
    }


def canonical_case(norm_key: str, variants: defaultdict[str, Counter[str]]) -> str:
    if norm_key not in variants:
        return norm_key
    return most_common_variant(variants[norm_key])


def auto_width(ws, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            max_len = max(max_len, len(str(val or "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, max_width)


def add_builder_rows(ws, start_row: int, group_name: str, keys: list[str], desc: dict[str, str], examples: dict[str, str]) -> int:
    row = start_row
    for key in keys:
        ws.cell(row=row, column=1, value=group_name)
        ws.cell(row=row, column=2, value=key)
        ws.cell(row=row, column=4, value=desc.get(key, ""))
        ws.cell(row=row, column=5, value=examples.get(key, ""))
        ws.cell(
            row=row,
            column=6,
            value=(
                f'=IF(C{row}="","",'
                f'IF(B{row}="spell",CHAR(9)&"spell"&CHAR(9)&C{row},'
                f'IF(LEFT(B{row},6)="CProp.",CHAR(9)&"CProp"&CHAR(9)&MID(B{row},7,200)&CHAR(9)&C{row},'
                f'CHAR(9)&B{row}&CHAR(9)&C{row})))'
            ),
        )
        row += 1
    return row


def generate_workbook(data: dict, spell_catalog: list[dict[str, str]], out_path: Path) -> Path:
    key_counts: Counter[str] = data["key_counts"]
    key_variants = data["key_variants"]
    key_pos_sum = data["key_pos_sum"]
    cprop_counts: Counter[str] = data["cprop_counts"]
    cprop_variants = data["cprop_variants"]
    cprop_pos_sum = data["cprop_pos_sum"]
    npc_spells: Counter[str] = data["npc_spells"]
    spell_variants = data["spell_variants"]

    descriptions = build_descriptions()

    key_rows = []
    for k_norm, count in key_counts.items():
        mean_pos = key_pos_sum[k_norm] / max(count, 1)
        case = canonical_case(k_norm, key_variants)
        key_rows.append((mean_pos, case, k_norm, count))
    key_rows.sort(key=lambda x: (x[0], x[1].lower()))

    cprop_rows = []
    for c_norm, count in cprop_counts.items():
        mean_pos = cprop_pos_sum[c_norm] / max(count, 1)
        case = canonical_case(c_norm, cprop_variants)
        cprop_rows.append((mean_pos, case, c_norm, count))
    cprop_rows.sort(key=lambda x: (x[0], x[1].lower()))

    # Build spell dropdown from: spell script names + npc-only spell tokens.
    spell_choices: dict[str, str] = {}

    for rec in spell_catalog:
        script = (rec.get("script") or "").strip()
        if script:
            spell_choices.setdefault(script.lower(), script)

    # Keep npc spell tokens that are not in spell catalog (e.g. custom/legacy aliases like massvatten).
    for s_norm in sorted(npc_spells.keys()):
        preferred = most_common_variant(spell_variants[s_norm])
        spell_choices.setdefault(s_norm, preferred)

    all_spells = sorted(spell_choices.values(), key=lambda v: v.lower())

    non_skill_keys = {
        "npctemplateid",
        "name", "script", "objtype", "color", "truecolor", "gender", "equip",
        "str", "int", "dex", "hits", "mana", "stam",
        "alignment", "hostile", "virtue", "karma", "fame",
        "lootgroup", "magicitemchance", "magicitemlevel",
        "cast_pct", "num_casts", "spell", "cprop",
        "attackattribute", "attackspeed", "attackdamage", "ar",
        "deathsnd", "tameskill", "food", "mount", "movemode", "runspeed",
        "dstart", "saywords", "speech", "provoke", "privs", "settings",
        "guardignore", "noloot", "title", "dress", "psub",
        "attackhitsound", "attackmisssound",
    }
    skill_candidates = [
        case for _, case, norm, _ in key_rows if norm not in non_skill_keys and norm != "cprop"
    ]

    wb = openpyxl.Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "Builder"

    title_fill = PatternFill("solid", fgColor="203864")
    header_fill = PatternFill("solid", fgColor="2F4F4F")
    title_font = Font(bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = "Builder Instructions"
    ws["B1"] = (
        "1) Fill inputs in column C.\n"
        "2) For skills, pick skill name in C and value in D (20 rows).\n"
        "3) For spells, pick spell names in the 20 spell rows.\n"
        "4) Copy the generated block from the Output sheet into npcdesc.cfg."
    )
    ws["B1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"] = "HP Guidance"
    ws["B2"] = (
        "STR is the base strength stat. CProp.BaseStrmod is an extra strength modifier applied by shard setup scripts after spawn.\n"
        "If HITS is filled, npcdesc can set a direct HP value from that field.\n"
        "If HITS is blank and CustomHitsLevel is not set, HP commonly follows the strength-based path, so STR plus BaseStrmod usually drives final max HP.\n"
        "If CProp.CustomHitsLevel exists, shard scripts treat it as the HP override and commonly read it in hundredths, so i30000 behaves like 300 HP.\n"
        "Use HITS for fixed HP, use STR plus BaseStrmod for stat-driven HP, and use CustomHitsLevel when you want an explicit override."
    )
    ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A3"] = "Template Id"
    ws["B3"] = "NpcTemplateID is the first row under Identity. Keep it short and do not include spaces."
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")

    for cell in ("A1", "A2", "A3"):
        ws[cell].font = header_font
        ws[cell].fill = header_fill

    for row_idx in (1, 2, 3):
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)

    headers = ["Group", "Key", "Value", "Skill Value", "What It Does", "Example", "Output Line"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    examples = {
        "NpcTemplateID": "etherealhorse",
        "Name": "a dragon",
        "script": "killpcs",
        "objtype": "0x3b",
        "Color": "0",
        "TrueColor": "0",
        "Gender": "0",
        "Equip": "dragon",
        "STR": "400",
        "INT": "200",
        "DEX": "120",
        "HITS": "400",
        "MANA": "200",
        "STAM": "120",
        "alignment": "evil",
        "hostile": "1",
        "virtue": "8",
        "Karma": "-10000",
        "Fame": "10000",
        "lootgroup": "133",
        "MagicItemChance": "90",
        "MagicItemLevel": "9",
        "spell": "flamestrike",
        "CProp.Type": "sDragonkin",
        "CProp.BaseStrmod": "i1000",
        "CProp.BaseIntmod": "i500",
        "CProp.BaseDexmod": "i100",
    }

    key_set = {row[1] for row in key_rows}
    cprop_set = {row[1] for row in cprop_rows}

    identity = [k for k in ["Name", "script", "objtype", "Color", "TrueColor", "Gender", "Equip"] if k in key_set]
    stats = [k for k in ["STR", "INT", "DEX", "HITS", "MANA", "STAM"] if k in key_set]
    combat = [k for k in ["AttackAttribute", "AttackSpeed", "AttackDamage", "AR"] if k in key_set]
    behavior = [k for k in ["alignment", "hostile", "virtue", "Karma", "Fame", "guardignore", "dstart", "speech", "saywords"] if k in key_set]
    loot = [k for k in ["lootgroup", "MagicItemChance", "MagicItemLevel"] if k in key_set]
    casting = [k for k in ["cast_pct", "num_casts", "spell"] if k in key_set]

    protection_cprops = [
        cp for cp in [
            "Type", "Permmr", "PermPoisonImmunity", "PermMagicImmunity", "PhysicalProtection",
            "FireProtection", "WaterProtection", "AirProtection", "EarthProtection",
            "NecroProtection", "HolyProtection",
        ] if cp in cprop_set
    ]
    scaling_cprops = [cp for cp in ["BaseStrmod", "BaseIntmod", "BaseDexmod", "CustomHitsLevel", "BaseHpRegen", "BaseManaRegen"] if cp in cprop_set]

    reserved = set(identity + stats + combat + behavior + loot + casting)
    skill_norms = {s.lower() for s in skill_candidates}
    remaining_keys = [row[1] for row in key_rows if row[1] not in reserved and row[1] != "CProp" and row[1].lower() not in skill_norms]

    cprop_reserved = set(protection_cprops + scaling_cprops)
    remaining_cprops = [row[1] for row in cprop_rows if row[1] not in cprop_reserved]

    row = 7

    def add_row(group: str, key: str, value_example: str = "") -> int:
        nonlocal row
        ws.cell(row=row, column=1, value=group)
        ws.cell(row=row, column=2, value=key)
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=descriptions.get(key, ""))
        ws.cell(row=row, column=6, value=value_example or examples.get(key, ""))
        ws.cell(
            row=row,
            column=7,
            value=(
                f'=IF(OR(B{row}="NpcTemplateID",C{row}=""),"",'
                f'IF(B{row}="spell",'
                f'"    spell"&REPT(" ",MAX(1,24-LEN("spell")))&C{row},'
                f'IF(LEFT(B{row},6)="CProp.",'
                f'"    CProp"&REPT(" ",MAX(1,24-LEN("CProp")))&MID(B{row},7,200)&REPT(" ",MAX(1,24-LEN(MID(B{row},7,200))))&C{row},'
                f'"    "&B{row}&REPT(" ",MAX(1,24-LEN(B{row})))&C{row})))'
            ),
        )
        row += 1
        return row - 1

    template_id_row = add_row("Identity", "NpcTemplateID")
    for k in identity:
        add_row("Identity", k)
    for k in stats:
        add_row("Primary Stats + Scaling", k)
    for c in scaling_cprops:
        add_row("Primary Stats + Scaling", f"CProp.{c}")
    for k in combat:
        add_row("Combat", k)
    for k in behavior:
        add_row("Behavior/Faction", k)
    for k in loot:
        add_row("Loot", k)

    for k in [k for k in casting if k != "spell"]:
        add_row("Casting + Spells", k)

    spell_rows: list[int] = []
    for _ in range(20):
        spell_rows.append(add_row("Casting + Spells", "spell", "flamestrike"))

    skill_rows: list[int] = []
    for _ in range(20):
        r = add_row("Skills", "skill_name", "Magery")
        skill_rows.append(r)
        ws.cell(r, 5, value=descriptions["skill_name"])
        ws.cell(r, 6, value="Magery")
        ws.cell(r, 4, value="100")
        ws.cell(r, 7, value=f'=IF(OR(C{r}="",D{r}=""),"","    "&C{r}&REPT(" ",MAX(1,24-LEN(C{r})))&D{r})')

    for c in protection_cprops:
        add_row("Protection CProps", f"CProp.{c}")
    for k in remaining_keys:
        add_row("Other Fields", k)
    for c in remaining_cprops:
        add_row("Other CProps", f"CProp.{c}")

    if spell_rows and all_spells:
        dv_sp = DataValidation(type="list", formula1=f"=SpellLists!$A$2:$A${len(all_spells)+1}", allow_blank=True)
        ws.add_data_validation(dv_sp)
        dv_sp.add(f"C{spell_rows[0]}:C{spell_rows[-1]}")

    if skill_rows and skill_candidates:
        dv_sk = DataValidation(type="list", formula1=f"=SkillLists!$A$2:$A${len(skill_candidates)+1}", allow_blank=True)
        ws.add_data_validation(dv_sk)
        dv_sk.add(f"C{skill_rows[0]}:C{skill_rows[-1]}")

    ws["H1"] = "NpcTemplate Output"
    ws["H1"].font = title_font
    ws["H1"].fill = title_fill
    ws["H2"] = f'="NpcTemplate "&SUBSTITUTE(TRIM(C{template_id_row})," ","")'
    ws["H3"] = "{"
    for out_row in range(4, 1004):
        src = out_row + 3
        ws.cell(out_row, 8, value=f'=IF($G{src}="","",$G{src})')
    ws["H1004"] = "}"

    ws.freeze_panes = "A7"
    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["H"].width = 80
    ws.row_dimensions[1].height = 70
    ws.row_dimensions[2].height = 120
    ws.row_dimensions[3].height = 36

    # --- FieldDictionary sheet ---
    ws_dict = wb.create_sheet("FieldDictionary")
    ws_dict.append(["Variable", "Type", "Description", "Observed Count", "Preferred Case"])
    for c in range(1, 6):
        cell = ws_dict.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill

    ws_dict.append([
        "Stats+Scaling precedence",
        "Note",
        "STR is the base stat, HITS is a direct hp assignment, BaseStrmod is an extra strength modifier, and CustomHitsLevel is the shard's hp override (commonly stored in hundredths, so i30000 becomes 300 hp).",
        "",
        "",
    ])

    for _, case, _k_norm, count in key_rows:
        var = case
        desc = descriptions.get(var, "Top-level NPC template field.")
        ws_dict.append([var, "TopLevel", desc, count, case])

    for _, case, _c_norm, count in cprop_rows:
        var = f"CProp.{case}"
        desc = descriptions.get(var, "Custom property attached to NPC via CProp.")
        ws_dict.append([var, "CProp", desc, count, case])

    ws_dict.append(["skill_name", "BuilderOnly", descriptions["skill_name"], 20, "skill_name"])
    ws_dict.append(["skill_value", "BuilderOnly", descriptions["skill_value"], 20, "skill_value"])
    ws_dict.append(["NpcTemplateID", "BuilderOnly", descriptions["NpcTemplateID"], 1, "NpcTemplateID"])

    # --- Output sheet ---
    ws_out = wb.create_sheet("Output")
    ws_out["A1"] = "NpcTemplate Output"
    ws_out["A1"].font = title_font
    ws_out["A1"].fill = title_fill
    ws_out["A2"] = f'="NpcTemplate "&SUBSTITUTE(TRIM(Builder!C{template_id_row})," ","")'
    ws_out["A3"] = "{"
    for out_row in range(4, 1004):
        src = out_row + 3
        ws_out.cell(out_row, 1, value=f'=IF(Builder!$G{src}="","",Builder!$G{src})')
    ws_out["A1004"] = "}"
    ws_out.column_dimensions["A"].width = 80

    # --- SpellLists sheet ---
    ws_sp = wb.create_sheet("SpellLists")
    ws_sp.append(["Spell", "InNpcdesc", "UseCountInNpcdesc"])
    for c in range(1, 4):
        cell = ws_sp.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill

    for sp in all_spells:
        s_norm = sp.lower()
        ws_sp.append([
            sp,
            "yes" if s_norm in npc_spells else "no",
            npc_spells.get(s_norm, 0),
        ])

    # --- SkillLists sheet ---
    ws_sk = wb.create_sheet("SkillLists")
    ws_sk.append(["Skill"])
    ws_sk.cell(1, 1).font = header_font
    ws_sk.cell(1, 1).fill = header_fill
    for sk in skill_candidates:
        ws_sk.append([sk])

    # --- AllVariables sheet ---
    ws_all = wb.create_sheet("AllVariables")
    ws_all.append(["Variable", "Category", "ObservedCount", "MeanOrderInTemplate", "PreferredCase"])
    for c in range(1, 6):
        cell = ws_all.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill

    for mean_pos, case, _k_norm, count in key_rows:
        ws_all.append([case, "TopLevel", count, round(mean_pos, 2), case])
    for mean_pos, case, _c_norm, count in cprop_rows:
        ws_all.append([f"CProp.{case}", "CProp", count, round(mean_pos, 2), case])

    for sheet in (ws, ws_dict, ws_out, ws_sp, ws_sk, ws_all):
        auto_width(sheet)

    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = out_path.with_name(f"{out_path.stem}_{stamp}{out_path.suffix}")
        wb.save(fallback)
        return fallback


def main() -> None:
    if not NPCDESC_PATH.exists():
        raise FileNotFoundError(f"Missing file: {NPCDESC_PATH}")

    data = parse_npcdesc(NPCDESC_PATH)
    spell_catalog = collect_spell_catalog(ROOT)
    saved_path = generate_workbook(data, spell_catalog, OUT_PATH)

    print(f"Saved {saved_path}")
    print(f"Templates parsed: {len(data['templates'])}")
    print(f"Top-level variables: {len(data['key_counts'])}")
    print(f"CProp variables: {len(data['cprop_counts'])}")
    print(f"NPC-used spells: {len(data['npc_spells'])}")
    print(f"Spell catalog entries: {len(spell_catalog)}")


if __name__ == "__main__":
    main()
